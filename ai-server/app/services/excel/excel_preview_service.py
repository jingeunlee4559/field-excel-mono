from pathlib import Path
from typing import Any

from fastapi import HTTPException
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries

from app.config import settings


def build_excel_preview(file_path: str | Path, sheet_name: str | None = None) -> dict[str, Any]:
    path = Path(file_path)

    if path.suffix.lower() not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        raise HTTPException(
            status_code=400,
            detail="엑셀 미리보기는 .xlsx, .xlsm 파일만 지원합니다.",
        )

    workbook = None

    try:
        workbook = load_workbook(path, data_only=True)
        sheet = workbook[sheet_name] if sheet_name else workbook.active

        bounds = get_used_bounds(sheet)

        merged_map, skip_cells = build_merged_cell_maps(sheet)

        columns = []
        column_widths: dict[str, int] = {}
        table_width = 0

        for col_idx in range(bounds["min_col"], bounds["max_col"] + 1):
            column_letter = get_column_letter(col_idx)
            excel_width = sheet.column_dimensions[column_letter].width or 10
            px_width = excel_width_to_px(excel_width)

            column_widths[column_letter] = px_width
            table_width += px_width
            columns.append({
                "index": col_idx,
                "letter": column_letter,
                "width": px_width,
            })

        rows = []

        for row_idx in range(bounds["min_row"], bounds["max_row"] + 1):
            excel_height = sheet.row_dimensions[row_idx].height or 24
            px_height = excel_height_to_px(excel_height)

            cells = []

            for col_idx in range(bounds["min_col"], bounds["max_col"] + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                coordinate = cell.coordinate

                if coordinate in skip_cells:
                    continue

                merge_info = merged_map.get(coordinate, {})
                rowspan = merge_info.get("rowspan", 1)
                colspan = merge_info.get("colspan", 1)

                width = 0

                for width_col_idx in range(col_idx, col_idx + colspan):
                    if width_col_idx > bounds["max_col"]:
                        continue

                    width += column_widths.get(get_column_letter(width_col_idx), 80)

                cells.append({
                    "coordinate": coordinate,
                    "address": coordinate,
                    "row": row_idx,
                    "col": col_idx,
                    "columnLetter": get_column_letter(col_idx),
                    "value": get_display_value(cell),
                    "rowspan": rowspan,
                    "colspan": colspan,
                    "width": width or column_widths.get(get_column_letter(col_idx), 80),
                    "height": px_height,
                    "style": get_cell_style(cell),
                })

            rows.append({
                "rowIndex": row_idx,
                "height": px_height,
                "cells": cells,
            })

        return {
            "template": {
                "fileName": path.name,
                "sheetName": sheet.title,
                "sheetNames": workbook.sheetnames,
            },
            "preview": {
                "sheetName": sheet.title,
                "bounds": bounds,
                "tableWidth": table_width,
                "columns": columns,
                "rows": rows,
            },
        }

    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(
            status_code=500,
            detail=f"엑셀 미리보기 생성 실패: {exc}",
        ) from exc
    finally:
        if workbook:
            workbook.close()


def get_used_bounds(sheet) -> dict[str, int]:
    min_row = sheet.max_row or 1
    max_row = 1
    min_col = sheet.max_column or 1
    max_col = 1

    has_value = False

    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is None:
                continue

            has_value = True
            min_row = min(min_row, cell.row)
            max_row = max(max_row, cell.row)
            min_col = min(min_col, cell.column)
            max_col = max(max_col, cell.column)

    for merged_range in sheet.merged_cells.ranges:
        min_col_m, min_row_m, max_col_m, max_row_m = range_boundaries(str(merged_range))
        min_row = min(min_row, min_row_m)
        max_row = max(max_row, max_row_m)
        min_col = min(min_col, min_col_m)
        max_col = max(max_col, max_col_m)

    if not has_value and not sheet.merged_cells.ranges:
        min_row = max_row = min_col = max_col = 1

    max_row = min(max_row, min_row + settings.MAX_EXCEL_ROWS - 1)
    max_col = min(max_col, min_col + settings.MAX_EXCEL_COLS - 1)

    return {
        "min_row": min_row,
        "max_row": max_row,
        "min_col": min_col,
        "max_col": max_col,
    }


def build_merged_cell_maps(sheet) -> tuple[dict[str, dict[str, int]], set[str]]:
    merged_map: dict[str, dict[str, int]] = {}
    skip_cells: set[str] = set()

    for merged_range in sheet.merged_cells.ranges:
        top_left = sheet.cell(merged_range.min_row, merged_range.min_col).coordinate

        merged_map[top_left] = {
            "rowspan": merged_range.max_row - merged_range.min_row + 1,
            "colspan": merged_range.max_col - merged_range.min_col + 1,
        }

        for row_idx in range(merged_range.min_row, merged_range.max_row + 1):
            for col_idx in range(merged_range.min_col, merged_range.max_col + 1):
                coordinate = sheet.cell(row_idx, col_idx).coordinate

                if coordinate != top_left:
                    skip_cells.add(coordinate)

    return merged_map, skip_cells


def get_display_value(cell) -> str:
    value = cell.value

    if value is None:
        return ""

    if isinstance(value, float) and value.is_integer():
        value = int(value)

    if isinstance(value, (int, float)):
        if cell.number_format and "#,##0" in cell.number_format:
            return f"{value:,.0f}"

        return str(value)

    return str(value)


def excel_width_to_px(width: float) -> int:
    return int(max(width * 8, 44))


def excel_height_to_px(height: float) -> int:
    return int(min(max(height * 1.35, 24), 620))


def color_to_hex(color) -> str:
    if not color:
        return ""

    try:
        if color.type == "rgb" and color.rgb:
            rgb = color.rgb

            if len(rgb) == 8:
                return f"#{rgb[2:]}"

            if len(rgb) == 6:
                return f"#{rgb}"

        if color.type == "indexed":
            return ""
    except Exception:
        return ""

    return ""


def border_style_to_css(border) -> str:
    if not border or not border.style:
        return "1px solid #e2e8f0"

    style_map = {
        "hair": "1px solid #e2e8f0",
        "thin": "1px solid #cbd5e1",
        "medium": "2px solid #94a3b8",
        "thick": "3px solid #64748b",
        "dashed": "1px dashed #94a3b8",
        "dotted": "1px dotted #94a3b8",
        "double": "3px double #94a3b8",
        "mediumDashed": "2px dashed #94a3b8",
        "dashDot": "1px dashed #94a3b8",
        "mediumDashDot": "2px dashed #94a3b8",
        "dashDotDot": "1px dotted #94a3b8",
        "mediumDashDotDot": "2px dotted #94a3b8",
    }

    return style_map.get(border.style, "1px solid #cbd5e1")


def get_cell_style(cell) -> dict[str, Any]:
    fill_color = ""

    try:
        if cell.fill and cell.fill.fill_type == "solid":
            fill_color = color_to_hex(cell.fill.fgColor)
    except Exception:
        fill_color = ""

    font_color = ""

    try:
        if cell.font and cell.font.color:
            font_color = color_to_hex(cell.font.color)
    except Exception:
        font_color = ""

    horizontal = cell.alignment.horizontal or "center"
    vertical = cell.alignment.vertical or "middle"
    wrap_text = bool(cell.alignment.wrap_text)
    font_size = float(cell.font.sz) if cell.font and cell.font.sz else 11
    font_weight = 800 if cell.font and cell.font.bold else 500

    return {
        # React 매핑 화면에서 바로 쓰는 CSS 키
        "backgroundColor": fill_color or "#ffffff",
        "color": font_color or "#0f172a",
        "fontWeight": font_weight,
        "fontSize": font_size,
        "textAlign": horizontal,
        "verticalAlign": vertical,
        "whiteSpace": "normal" if wrap_text else "nowrap",
        "borderTop": border_style_to_css(cell.border.top),
        "borderRight": border_style_to_css(cell.border.right),
        "borderBottom": border_style_to_css(cell.border.bottom),
        "borderLeft": border_style_to_css(cell.border.left),

        # 디버깅/확장용 원본 성격 키
        "fontBold": bool(cell.font.bold) if cell.font else False,
        "fontItalic": bool(cell.font.italic) if cell.font else False,
        "fontColor": font_color or "#0f172a",
        "horizontalAlign": horizontal,
        "verticalAlignRaw": vertical,
        "wrapText": wrap_text,
        "numberFormat": cell.number_format or "",
    }
